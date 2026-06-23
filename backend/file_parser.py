import io
import fitz  # PyMuPDF
from docx import Document
import openpyxl
import re
import statistics


# Sections that are NOT requirement modules — never emit [MODULE: ] for these.
# Covers: introductory, administrative, and reference sections found in most SRS documents.
_NON_REQ_SECTIONS = re.compile(
    r'''(?xi)
    ^(?:
        \d+[\.]?\s*                          # optional leading number like "1." "1.2"
    )?(?:
        introduction | scope | purpose | overview | background | foreword | preface |
        summary | executive\s+summary |
        references? | normative\s+references? | informative\s+references? |
        applicable\s+documents? | related\s+documents? |
        definitions? | abbreviations? | acronyms? | glossary | terms? |
        table\s+of\s+contents? | contents? | index |
        revision\s+history | change\s+(log|history|record) | document\s+history |
        deleted | obsolete | tbd | tbc | reserved | placeholder |
        document\s+control | document\s+organization | document\s+structure |
        applicability | general\s+information | general |
        list\s+of\s+(figures?|tables?) |
        nomenclature | standards?
    )\s*$
    ''',
    re.IGNORECASE | re.VERBOSE,
)


def parse_pdf(file_bytes: bytes) -> str:
    """
    Improved PDF parser that mirrors parse_docx behaviour:
    - Detects section headings via font size and bold formatting
    - Injects ## and [MODULE: ...] markers so document_ingestion.py
      can correctly identify module boundaries
    - SKIPS common non-requirement sections (Introduction, Scope, References…)
      so they never pollute the module list or steal pending_module from
      real requirement sections.

    Strategy
    --------
    Pass 1  Collect every font size in the document to find the median
            body-text size.
    Pass 2  For each text line:
            - If its max span size > body_size * 1.15  →  heading candidate
            - If any span is bold AND line is short (<= 100 chars) →  heading candidate
            - Heading candidates are then filtered:
                * Lines matching a requirement-ID pattern  →  NOT a heading
                * Lines matching _NON_REQ_SECTIONS         →  NOT a heading
            Surviving headings get  ## prefix  +  [MODULE: ...] marker.
            Everything else is emitted as plain text.
    """
    import statistics as _stats

    _REQ_LINE = re.compile(
        r'^[A-Z][A-Z0-9_\-\.]*\d+\s+(?:shall|must|should|will|the|is|are|a\s)',
        re.IGNORECASE,
    )

    parts: list = []
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")

        # ── Pass 1: determine body-text font size ─────────────────────────────
        all_sizes: list = []
        for page in doc:
            for block in page.get_text("dict")["blocks"]:
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        txt = span.get("text", "").strip()
                        sz  = span.get("size", 0)
                        if txt and sz > 0:
                            all_sizes.append(round(sz, 1))

        body_size        = _stats.median(all_sizes) if all_sizes else 10.0
        heading_min_size = body_size * 1.15          # >15 % larger → heading

        # ── Pass 2: extract text with structure markers ───────────────────────
        for page in doc:
            for block in page.get_text("dict")["blocks"]:
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    spans = line.get("spans", [])
                    if not spans:
                        continue

                    line_text = "".join(s.get("text", "") for s in spans).strip()
                    if not line_text or len(line_text) < 2:
                        continue

                    max_size = max((s.get("size", 0) for s in spans), default=0)
                    is_bold  = any(
                        "bold" in s.get("font", "").lower()
                        for s in spans if s.get("text", "").strip()
                    )

                    # Is this line a heading?
                    larger_font  = max_size > heading_min_size
                    bold_heading = is_bold and len(line_text) <= 100
                    is_heading   = larger_font or bold_heading

                    # Never treat a requirement-ID line as a heading
                    # Disqualify headings that are requirement ID lines
                    if is_heading and _REQ_LINE.match(line_text):
                        is_heading = False

                    # Disqualify headings that are non-requirement admin sections
                    # Strip leading section numbers before checking (e.g. "1. Introduction")
                    _bare = re.sub(r'^\d+(?:\.\d+)*[.:\s]+', '', line_text).strip()
                    if is_heading and _NON_REQ_SECTIONS.match(_bare):
                        is_heading = False

                    if is_heading:
                        parts.append(f"\n## {line_text}")
                        parts.append(f"[MODULE: {line_text}]")
                    else:
                        parts.append(line_text)

        doc.close()
    except Exception as e:
        raise RuntimeError(f"PDF parsing error: {e}")
    return "\n".join(parts)


def _paragraph_is_bold_heading(para) -> bool:
    """
    Returns True if an entire paragraph is bold-formatted (module heading).
    Checks both paragraph-level bold runs and character-level formatting.
    """
    runs = [r for r in para.runs if r.text.strip()]
    if not runs:
        return False
    # All non-empty runs must be bold
    return all(r.bold for r in runs)


def parse_docx(file_bytes: bytes) -> str:
    """
    Parses a DOCX file, preserving:
    - Bold headings as [MODULE: <text>] markers for module detection
    - Normal style headings as ## markers
    - Tables with pipe-separated rows
    Requirement: per spec §5.3, bold-formatted headings define module boundaries.
    """
    parts = []
    try:
        doc = Document(io.BytesIO(file_bytes))
        # Words that appear as heading text but are NOT real module names.
        _SKIP_HEADING_WORDS = {
            "deleted", "obsolete", "reserved", "tbd", "tbc", "na",
            "none", "placeholder", "empty", "blank", "unknown",
        }

        # Track heading hierarchy so skipped headings can fall back to parent
        _heading_stack = {}  # level (1-6) -> last seen heading text
        import re as _re

        def _emit_para(para):
            text = para.text.strip()
            if not text:
                return
            style_name = para.style.name if para.style else ""
            is_heading = style_name.startswith("Heading") or _paragraph_is_bold_heading(para)
            if is_heading:
                parts.append(f"\n## {text}")
                _lm = _re.match(r"Heading (\d+)", style_name)
                _level = int(_lm.group(1)) if _lm else 99
                if text.strip().lower() not in _SKIP_HEADING_WORDS:
                    parts.append(f"[MODULE: {text}]")
                    _heading_stack[_level] = text
                    for _l in list(_heading_stack):
                        if _l > _level: del _heading_stack[_l]
                else:
                    # Skipped heading — re-emit nearest valid parent as MODULE
                    _parent = next(
                        (_heading_stack[_l] for _l in sorted(_heading_stack) if _l < _level),
                        None
                    )
                    if _parent:
                        parts.append(f"[MODULE: {_parent}]")
            else:
                parts.append(text)

        def _emit_table(table):
            for row in table.rows:
                row_text = " | ".join(
                    cell.text.strip() for cell in row.cells if cell.text.strip()
                )
                if row_text:
                    parts.append(row_text)

        # Walk document body IN ORDER (paragraphs and tables interleaved).
        # This is critical: table requirements must inherit the heading that
        # immediately precedes their table in the document, not the last
        # heading anywhere in the file.
        from docx.text.paragraph import Paragraph as _Para
        from docx.table import Table as _Table
        for _child in doc.element.body:
            _tag = _child.tag.split("}")[-1] if "}" in _child.tag else _child.tag
            if _tag == "p":
                _emit_para(_Para(_child, doc))
            elif _tag == "tbl":
                _emit_table(_Table(_child, doc))
    except Exception as e:
        raise RuntimeError(f"DOCX parsing error: {e}")
    return "\n".join(parts)


def parse_xlsx(file_bytes: bytes) -> str:
    parts = []
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"\n[Sheet: {sheet_name}]")
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if all(v is None for v in row):
                    continue
                if i == 0:
                    headers = [str(c).strip() if c is not None else f"Col{j}" for j, c in enumerate(row)]
                    continue
                for h, val in zip(headers, row):
                    if val is not None and str(val).strip():
                        parts.append(f"{h}: {val}")
        wb.close()
    except Exception as e:
        raise RuntimeError(f"Excel parsing error: {e}")
    return "\n".join(parts)


def parse_file(filename: str, file_bytes: bytes) -> str:
    if not filename or not file_bytes:
        raise ValueError("Filename and file content are required")
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "pdf":
        return parse_pdf(file_bytes)
    elif ext in ("docx", "doc"):
        return parse_docx(file_bytes)
    elif ext in ("xlsx", "xls"):
        return parse_xlsx(file_bytes)
    else:
        raise ValueError(
            f"Unsupported file type: .{ext}. Accepted: .pdf, .docx, .xlsx"
        )